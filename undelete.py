import os
import sys
from struct import *
from datetime import datetime
from typing import Iterable
import argparse
import openpyxl
from typing import List

DATA_START_BYTE = 0
CLUSTER_SIZE = 0
RESERVED_SECTORS = 0
BYTES_PER_SECTOR = 0
FATSize = 0
TOTAL_SECTORS = 0
TOTAL_CLUSTERS = 0
# 有効なクラスタ番号を抽出するビットマスク
FAT32_CLUSTER_MASK = 0x0FFFFFFF


def lookup_path(excel_file_path: str):
    """
    Excelファイルから、親ディレクトリ特定用の辞書を作成する
    Key: ディレクトリの先頭クラスタ番号 (int)
    Value: そのディレクトリの全情報 (dict)
    """
    wb = openpyxl.load_workbook(excel_file_path)
    ws = wb.active
    
    parent_lookup = {}
    
    # １周目（ディレクトリ辞書の作成）
    # ヘッダー行をスキップ (min_row=2)
    for row in ws.iter_rows(min_row=2):
        
        # 必要な列の値を抽出 (値は row[index].value で取得)
        # 属性: Index 6, 先頭クラスタ: Index 8, クラスタ位置: Index 2
        
        attribute = row[6].value
        first_cluster = row[8].value
        delete_frag = row[9].value
        
        # 🚨 最重要: ディレクトリのエントリのみを候補として辞書に登録する
        # (ファイルエントリは親にはなれない。LFNエントリはスキップ)
        if attribute == '0x10' and delete_frag != '!':  # ディレクトリ属性で削除フラグが立っていない
            
            # 先頭クラスタをキーとし、そのディレクトリの情報を値として格納
            # ここではシンプルに、全ての列をタプルとして値にする
            # 実際にはタプルではなく、辞書に変換した方が扱いやすいです
            parent_lookup[int(first_cluster)] = {
                "filename": row[3].value,
                "cluster_location": int(row[2].value),
                "entry_row_data": [cell.value for cell in row]
            }

    # ２周目（パスの特定と書き込み）
    # ヘッダー行をスキップ (min_row=2)
    for row in ws.iter_rows(min_row=2):
        # 現在チェックしているファイルの「格納場所（クラスタ位置）」
        current_location_cluster = int(row[2].value)
        attribute = row[6].value
        # フルパスを保存するリスト（最初は自分自身のファイル名）
        path_list = []
        while True:
            # 辞書から親ディレクトリを探す
            # parent_lookup のキーは「先頭クラスタ」= 子ファイルの「格納場所」
            parent_entry = parent_lookup.get(current_location_cluster)

            if parent_entry:
                # 親が見つかった場合
                parent_name = parent_entry['filename']
                path_list.insert(0, parent_name) # リストの先頭に追加
                
                # 次の検索のために、キーを「親の格納場所」に更新してループ継続
                current_location_cluster = int(parent_entry['cluster_location'])
                
                # ※無限ループ防止（ルートディレクトリ自身を指している場合などの対策）
                if current_location_cluster == 0 or current_location_cluster == 2:
                    path_list.insert(0, "ROOT")
                    break
            else:
                # 親が見つからなかった場合（ルートに到達 or 孤立）
                path_list.insert(0, "ROOT")
                break

        # リストを繋げてパスにする
        full_path = "\\".join(path_list)
        if full_path != "ROOT":
            print(f"復元パス: {full_path}")
        row[10].value = full_path
    
    # 3. ファイルの保存
    try:
        wb.save(excel_file_path)
        print(f"\n✅ Excelファイルに保存が完了しました: {excel_file_path}")
    except Exception as e:
        print(f"\n❌ ファイル保存中にエラーが発生しました: {e}")

def salvage_file(excel_file_path: str):
    global DATA_START_BYTE, CLUSTER_SIZE, RESERVED_SECTORS, BYTES_PER_SECTOR
    """
    Excelファイルの「復旧チェック」列を見て、「1」の時、復旧する
    excel_file_path: エクセルファイル名
    """
    wb = openpyxl.load_workbook(excel_file_path)
    ws = wb.active

    for row in ws.iter_rows(min_row=2):
        file_full_path = "\\".join([row[10].value, row[3].value])
        parent_dir = row[10].value
        file_name = row[3].value
        
        if row[0].value != 1: # 復元対象外
            continue

        # 復元対象
        first_cluster = int(row[8].value)
        file_size = int(row[5].value)
        CLUSTER_SIZE = int(row[14].value)
        RESERVED_SECTORS = int(row[11].value)
        BYTES_PER_SECTOR = int(row[12].value)
        DATA_START_BYTE = int(row[15].value)
        # 親ディレクトリを作成
        os.makedirs(parent_dir, exist_ok=True)
        
        file_size_rest = file_size
        cluster_chain = [first_cluster]
        current_cluster = first_cluster
        while file_size_rest > 0:
            current_cluster = get_next_cluster("D", current_cluster)
            cluster_chain.append(current_cluster)
            file_size_rest -= CLUSTER_SIZE
        # 最後の要素は削除
        cluster_chain.pop()
        # print(f"cluster chain: {cluster_chain}")
        
        file_size_rest = file_size
        file_data = b""
        for cluster in cluster_chain: # クラスタチェーンをたどって、１クラスタずつ読む
            file_data += get_file("D", cluster, file_size_rest if file_size_rest < CLUSTER_SIZE else CLUSTER_SIZE)
            file_size_rest -= CLUSTER_SIZE
        with open(file_full_path, "wb") as out_f:
            out_f.write(file_data)

        # 指定した日時に変更する
        update_datetime = datetime.strptime(row[7].value, "%Y-%m-%d %H:%M:%S")
        os.utime(path=file_full_path, times=(update_datetime.timestamp(), update_datetime.timestamp()))
        # 復旧チェック列を「０」にする
        row[0].value = 0
        print(f"ファイルを保存しました: {file_full_path}")

    # 3. ファイルの保存
    try:
        wb.save(excel_file_path)
        print(f"\n✅ Excelファイルに保存が完了しました: {excel_file_path}")
    except Exception as e:
        print(f"\n❌ ファイル保存中にエラーが発生しました: {e}")


# 辞書を作成
# parent_map = create_parent_lookup("fat32_scan_results.xlsx")
def sanitize_string(value: str, invalid_codepoints: Iterable[int] | None = None) -> str:
    """
    文字列中の制御コードや不正な文字を除去する。

    Parameters
    ----------
    value: str
        対象文字列
    invalid_codepoints: Iterable[int] | None
        除去したいUnicodeコードポイントのリスト。省略時は0x00〜0x1Fと0x7F。
    """
    if invalid_codepoints is None:
        invalid_codepoints = list(range(0x20)) + [0x7F]
    invalid_chars = {chr(cp) for cp in invalid_codepoints}
    # translateを使って高速に削除
    translation_table = {ord(ch): None for ch in invalid_chars}
    return value.translate(translation_table)

def save_to_excel(results: list, output_filename: str = "fat32_scan_results.xlsx"):

    # 新しいワークブックを作成
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "復旧可能エントリリスト"
    
    # 1. ヘッダー行の書き込み
    headers = ["復旧チェック", "バイト位置", "クラスタ位置", "ファイル名", "ファイルタイプ", "ファイルサイズ (B)", "属性", "最終更新日時", "先頭クラスタ", "削除フラグ", "場所", "RESERVED_SECTORS", "BYTES_PER_SECTOR", "FATSize", "CLUSTER_SIZE", "DATA_START_BYTE"]
    #           A               B             C               D             E                 F                     G       H               I               J             K       L                   M                   N          O               P
    ws.append(headers)
    
    # 2. データ行の書き込み
    for entry in results:
        # Excelの1行に書き込むデータ
        row_data = [
            "",
            entry["current_byte"],
            entry["current_cluster"],
            entry["filename"],                   # ファイル名 (LFN または SFN)
            entry["filetype"],                   # ファイルタイプ（８．３ファイル名の拡張子）
            entry["size"],                       # ファイルサイズ
            entry["attribute"],                    # ファイル属性
            entry["updatetime"],                    # 日時情報 (文字列または datetime オブジェクト)
            entry["first_cluster"],                    # 先頭クラスタ番号
            "!" if entry["deleted"] else "", # 削除フラグ
            "", # 場所（２周目で算出）
            RESERVED_SECTORS,
            BYTES_PER_SECTOR,
            FATSize,
            CLUSTER_SIZE,
            DATA_START_BYTE
        ]
        ws.append(row_data)
        ws[f'B{ws.max_row}'].number_format = '#,##0'
        ws[f'C{ws.max_row}'].number_format = '#,##0'
        ws[f'F{ws.max_row}'].number_format = '#,##0'
        ws[f'I{ws.max_row}'].number_format = '#,##0'
        ws[f'L{ws.max_row}'].number_format = '#,##0'
        ws[f'M{ws.max_row}'].number_format = '#,##0'
        ws[f'N{ws.max_row}'].number_format = '#,##0'
        ws[f'O{ws.max_row}'].number_format = '#,##0'
        ws[f'P{ws.max_row}'].number_format = '#,##0'
        
    # 3. ファイルの保存
    try:
        wb.save(output_filename)
        print(f"\n✅ Excelファイルに保存が完了しました: {output_filename}")
    except Exception as e:
        print(f"\n❌ ファイル保存中にエラーが発生しました: {e}")

def get_file(drive_letter: str, first_cluster: int, file_size: int) -> bytes:
    global DATA_START_BYTE
    """
    FAT32のクラスタ番号からファイルデータを読み込む関数
    """
    drive_path = f"\\\\.\\{drive_letter}:"
    if first_cluster < 2:
        # クラスタ0と1は予約済み
        raise ValueError(
            "無効なクラスタ番号です。クラスタ番号は2以上である必要があります。"
        )

    # クラスタ2がデータ領域の先頭（オフセット0）に対応する
    offset = DATA_START_BYTE + (first_cluster - 2) * CLUSTER_SIZE

    with open(drive_path, "rb") as f:
        f.seek(offset)
        return f.read(file_size)

def get_next_cluster(drive_letter: str, cluster_number: int) -> int:
    global RESERVED_SECTORS, BYTES_PER_SECTOR
    """
    次のクラスタ番号を返す関数
    """
    # Windowsでは「\\.\」を前につけて特殊なデバイスとして扱う必要があるわ。
    drive_path = f"\\\\.\\{drive_letter}:"
    if cluster_number < 2:
        # クラスタ0と1は予約済み
        raise ValueError(
            "無効なクラスタ番号です。クラスタ番号は2以上である必要があります。"
        )

    with open(drive_path, "rb") as f:
        # 読み込むデータサイズ (1MB = 1024 * 1024 バイト)
        READ_SIZE = 1024 * 1024
        raw_data = f.read(READ_SIZE)
        # offset = RESERVED_SECTORS * BYTES_PER_SECTOR + cluster_number * 4
        offset = RESERVED_SECTORS * BYTES_PER_SECTOR - 512
        f.seek(offset)
        f.read(512 + cluster_number * 4)
        next_cluster = unpack("<I", f.read(4))[0]
        # 6. FAT32の予約ビットをクリア (上位4ビットを無視)
        next_cluster &= FAT32_CLUSTER_MASK
        return next_cluster


def read_raw_data(drive_letter: str, target_exts: List[str], xlsx_file: str):
    # Windowsでは「\\.\」を前につけて特殊なデバイスとして扱う必要があるわ。
    drive_path = f"\\\\.\\{drive_letter}:"
    # 読み込むデータサイズ (1MB = 1024 * 1024 バイト)
    READ_SIZE = 1024 * 1024
    # 復元対象としたい属性のマスク
    VALID_ATTRIBUTES = 0x08 | 0x10 | 0x20  # 0x08(V) or 0x10(D) or 0x20(A)
    
    try:
        # バイナリ読み込みモード ('rb') でドライブを開く
        print(f"[{drive_path}] の生データ読み込みを開始します...")

        with open(drive_path, "rb") as f:
            # データを読み込む
            raw_data = f.read(READ_SIZE)

            # 読み込んだデータサイズを確認
            actual_size = len(raw_data)
            print(f"成功: {actual_size} バイトのデータを読み込みました。")

            ## --- ここから読み込んだデータの確認 --- ##

            # 1. データ全体の先頭16バイトを16進数で表示
            print("\n--- 先頭16バイト (16進数) ---")
            # b'' 形式で出力されるのを避けるために .hex() を使うと綺麗よ
            print(raw_data[:16].hex())

            # 2. 最初の512バイト（ブートセクタ）を別のファイルに保存して確認するのもアリ！
            # with open("D_boot_sector.bin", "wb") as out_f:
            #     out_f.write(raw_data[:512])
            # print("\n最初の512バイトを [D_boot_sector.bin] に保存しました。")

            # 3. FAT32の署名（510-511バイト目）を確認
            # FATのブートセクタの最後の2バイトは必ず 0x55AA になっているはずよ！
            boot_signature = raw_data[510:512]
            print("\n--- ブートシグネチャ (510-511バイト目) ---")
            # リトルエンディアンで 'AA 55' と表示されるはずよ
            print(boot_signature.hex())

            if boot_signature == b"\x55\xaa":
                print("✔ ブートシグネチャ [0x55AA] を確認！これは有効なブートセクタよ。")
                BYTES_PER_SECTOR = unpack("<H", raw_data[11:13])[0]
                print(f"BYTES_PER_SECTOR: {BYTES_PER_SECTOR}")
                SectorsPerCluster = unpack("<B", raw_data[13:14])[0]
                print(f"SectorsPerCluster: {SectorsPerCluster}")
                RESERVED_SECTORS = unpack("<H", raw_data[14:16])[0]
                print(f"RESERVED_SECTORS: {RESERVED_SECTORS}")
                FATCount = unpack("<B", raw_data[16:17])[0]
                print(f"FATCount: {FATCount}")
                RootDirectoryEntries = unpack("<I", raw_data[44:48])[0]
                print(f"RootDirectoryEntries: {RootDirectoryEntries}")
                FATSize = unpack("<I", raw_data[36:40])[0]
                print(f"FATSize: {FATSize}")
                TOTAL_SECTORS = unpack("<I", raw_data[32:36])[0]
                print(f"TOTAL_SECTORS: {TOTAL_SECTORS}")
                # RootDirectoryCluster = unpack("<H", raw_data[21:23])
                # print(f"RootDirectoryCluster: {RootDirectoryCluster}")
                # FSInfoSector = unpack("<H", raw_data[23:25])
                # print(f"FSInfoSector: {FSInfoSector}")
                # BackupBootSector = unpack("<H", raw_data[25:27])
                # print(f"BackupBootSector: {BackupBootSector}")
                # BackupDataSector = unpack("<H", raw_data[27:29])
                # print(f"BackupDataSector: {BackupDataSector}")

                data_start_sector = RESERVED_SECTORS + (FATCount * FATSize)
                print(f"data_start_sector: {data_start_sector}")

                DATA_START_BYTE = data_start_sector * BYTES_PER_SECTOR
                print(f"DATA_START_BYTE: {DATA_START_BYTE}")

                CLUSTER_SIZE = SectorsPerCluster * BYTES_PER_SECTOR
                print(f"CLUSTER_SIZE: {CLUSTER_SIZE}")
                
                TOTAL_CLUSTERS = (TOTAL_SECTORS * BYTES_PER_SECTOR - DATA_START_BYTE) // CLUSTER_SIZE
                print(f"TOTAL_SECTORS: {TOTAL_CLUSTERS}")

                f.seek(DATA_START_BYTE)
                byte_read = 0
                lfn_buffer = []
                scan_results = []
                deleted_file = False
                while True: # byte_read < BYTES_PER_SECTOR * 1024 * 1024 * 2 * 1:
                    deleted_file = False
                    data = f.read(32)

                    # ファイルの終端に達したら、f.read() は空のバイト列 (b'') を返す
                    if not data:
                        print("ドライブの物理的な終端に到達しました。スキャンを終了します。")
                        break
                    
                    # 💡 データが32バイト未満の場合 (ドライブの終端が32の倍数でない場合)
                    if len(data) < 32:
                        print(f"終端で {len(data)} バイトを読み込みました。スキャンを終了します。")
                        break

                    byte_read += 32

                    (
                        filename_bytes,
                        extension_bytes,
                        attribute_byte,
                        first_cluster,
                        file_size,
                    ) = unpack("<8s 3s B L L", data[:12] + data[26:28] + data[20:22] + data[28:32])

                    # 読み込んだ first_cluster にマスクを適用
                    first_cluster = first_cluster & FAT32_CLUSTER_MASK
                    # 先頭クラスタ位置がトータルクラスタよりも大きければ無効
                    # if first_cluster < 2 or first_cluster > TOTAL_CLUSTERS:
                    #     continue
                    # --- LFN (0x0F)とその他の属性のチェック ---
                    if (attribute_byte & 0x0F) == 0x0F:
                        # sequence_number = (
                        #     data[0] & 0x3F
                        # )  # 0x3F (63) でマスクして上位ビット（0x40）を無視
                        # LFNエントリから名前の断片（13文字分）とチェックサムを取得
                        lfn_name_part_bytes = data[1:11] + data[14:26] + data[28:32]
                        # バッファに保存（シーケンス番号、バイト列、チェックサム）
                        lfn_buffer.append(
                            {
                                "seq": data[0],
                                "bytes": lfn_name_part_bytes,
                                "checksum": data[13],
                                # ... 他の情報も保存
                            }
                        )
                        continue  # 次の32バイトへ

                    # 属性が 0x00 の場合、空のエントリか未使用の場所なのでスキップ
                    if attribute_byte == 0x00:
                        # 0x00を見つけたら、それ以降は未使用の可能性が高いから、セクタの残りもスキップして次のセクタへ移動してもいいくらいよ。
                        # print("  -> 未使用エントリをスキップ")
                        # 一応LFNエントリをクリア
                        lfn_buffer = []
                        continue  # 次の32バイトへ

                    # 属性が有効なファイル・ディレクトリであることを確認
                    # ここでは、属性値が 0x0F (LFN) を含まない、有効な属性ビットのいずれかを持っているか確認
                    if attribute_byte & VALID_ATTRIBUTES and not (attribute_byte & 0x0F):
                        # 有効なファイルまたはディレクトリのエントリである可能性が高い！
                        # 削除ファイルか？
                        if data[0:1] == b"\xe5":
                            filename_list = list(filename_bytes)
                            filename_list[0] = 0x21 # 「!」に置換
                            filename_bytes = bytes(filename_list)
                            deleted_file = True

                        filename_str = sanitize_string(
                            filename_bytes.decode("shift_jis", errors="ignore")
                        ).strip()
                        extension_str = sanitize_string(
                            extension_bytes.decode("shift_jis", errors="ignore")
                        ).strip()

                        # if filename_str != "" and extension_str != "":
                            # if extension_str in ["DOC", "XLS", "JPG", "PDF", "PNG", "PPT", "PAG"] and (attribute_byte & 0x20) != 0 or ((attribute_byte & 0x10) != 0 and file_size == 0):
                        if (hex(attribute_byte >> 4) == '0x2' and file_size > 0) or (hex(attribute_byte >> 4) == '0x1' and file_size == 0):
                            full_filename_str = f"{filename_str}.{extension_str}"
                            # LFNを持っているか
                            if lfn_buffer:
                                # 最初のシーケンス番号が0x40以上か and 最初のシーケンス番号が要素の数と一致しているか？
                                if lfn_buffer[0]["seq"] >= 0x40 and lfn_buffer[0]["seq"] & 0x3F == len(lfn_buffer):
                                    # 3. LFNシーケンスをシーケンス番号順にソート（逆順に保存されているため）
                                    lfn_buffer.sort(key=lambda x: x["seq"])
                                    # 4. バイト列を結合して、UTF-16でデコード
                                    full_name_bytes = b""
                                    for part in lfn_buffer:
                                        full_name_bytes += bytes(part["bytes"])
                                    # UTF-16LE (リトルエンディアン) でデコードし、終端の \x00 を取り除く
                                    try:
                                        decoded_full_name = full_name_bytes.decode("utf-16le")
                                        full_filename_str = sanitize_string(
                                            decoded_full_name.split("\x00", 1)[0]
                                        )
                                    except Exception as e:
                                        lfn_buffer = []
                                        continue
                                lfn_buffer = []

                            # 更新日
                            # とりあえず現在時刻をセット
                            update_datetime = datetime.now()
                            date_value = unpack("<H", data[24:26])[0]
                            # 1. 日（Bit 4～0）の抽出
                            day = (
                                date_value & 0x1F
                            )  # 0x1F は 0b00011111 (5ビットすべて1)
                            # 2. 月（Bit 8～5）の抽出
                            month = (
                                date_value >> 5
                            ) & 0x0F  # 0x0F は 0b00001111 (4ビットすべて1)
                            # 3. 年（Bit 15～9）の抽出
                            # 0x7F は 0b01111111 (7ビットすべて1)
                            year_offset = (date_value >> 9) & 0x7F
                            actual_year = 1980 + year_offset

                            # 更新時間
                            time_value = unpack("<H", data[22:24])[0]
                            # 1. 秒（Bit 4～0）の抽出
                            second = (
                                time_value & 0x1F
                            ) * 2  # 0x1F は 0b00011111 (5ビットすべて1)
                            # 2. 分（Bit 10～5）の抽出
                            minute = (
                                time_value >> 5
                            ) & 0b00111111  # 0x3F は 0b00111111 (6ビットすべて1)
                            # 3. 時（Bit 15～11）の抽出
                            # 0x1F は 0b00011111 (5ビットすべて1)
                            hour = (time_value >> 11) & 0x1F

                            try:
                                update_datetime = datetime(actual_year, month, day, hour, minute, second)
                            except Exception as e:
                                # 有効な日付でなければ抜ける
                                continue
                            
                            # ディレクトリ or 対象拡張子でなければ抜ける
                            if not (extension_str in target_exts  or full_filename_str[-6:] == '.pages' or (hex(attribute_byte >> 4) == '0x1' and file_size == 0)):
                                # if not (extension_str in ["DOC", "XLS", "JPG", "PDF", "PNG", "PPT", "PAG"]  or full_filename_str[-6:] == '.pages'):
                                #     print(f"対象外ファイルタイプ : {extension_str} - {full_filename_str[-6:]}")
                                continue
                            # ディレクトリ and ファイル名が「」か「..」か「.」なら抜ける
                            if hex(attribute_byte >> 4) == '0x1' and file_size == 0 and (filename_str == "" or filename_str == ".." or filename_str == "."):
                                continue
                            # 先頭クラスタ位置がトータルクラスタよりも大きければ無効
                            if (first_cluster < 2 or first_cluster > TOTAL_CLUSTERS):
                                pass
                                continue

                            current_byte = f.tell()
                            current_cluster = (current_byte - DATA_START_BYTE) // CLUSTER_SIZE  + 2
                            scan_results.append({
                                "current_byte": current_byte,
                                "current_cluster": current_cluster,
                                "filename": full_filename_str,
                                "filetype": extension_str,
                                "size": file_size,
                                "attribute": hex(attribute_byte),
                                "updatetime": update_datetime.strftime("%Y-%m-%d %H:%M:%S"),
                                "first_cluster": first_cluster,
                                "deleted": deleted_file
                            })
                            
                            print(f"{current_cluster}/{TOTAL_CLUSTERS}  ファイルエントリ->{full_filename_str} --- {extension_str} --- {hex(attribute_byte)} --- {file_size} bytes --- {first_cluster}")
                # whileループ脱出後
                # excelに保存
                save_to_excel(scan_results, xlsx_file)

            else:
                print("⚠ ブートシグネチャが見つかりません。ドライブへのアクセスに問題があるかも...")

    except FileNotFoundError:
        print(f"エラー: ドライブ [{drive_path}] が見つかりません。接続を確認してね。")
    except PermissionError:
        print(
            f"エラー: ドライブ [{drive_path}] へのアクセス権限がありません。管理者権限で実行する必要があるかも。"
        )
    except Exception as e:
        print(f"予期せぬエラーが発生しました: {e}")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="FAT32ドライブからデータを復旧"
    )

    parser.add_argument("--target_drive", "-t", type=str, required=True, help="復旧対象ドライブレター（a,b,c,...）")
    
    run_mode = parser.add_mutually_exclusive_group(required=True)  # 同グループ内のどれか必須
    run_mode.add_argument("--scan", "-s", action="store_true", help="スキャンモード実行（スキャン結果をエクセルファイルに保存）")
    run_mode.add_argument("--restore", "-r", action="store_true", help="復元モードで実行（スキャン結果ファイル中の復旧フラグファイルを復元）")

    parser.add_argument(
        "--extensions", "-e",
        nargs="+",
        required=False,
        default=["DOC", "XLS", "JPG", "PDF", "PNG", "PPT", "PAG"] ,
        help="復旧対象とする拡張子（スペースで区切って複数指定可）"
    )
    parser.add_argument("--xlsx_file", "-x", type=str, required=False, help="復旧対象ドライブレター（a,b,c,...）", default='fat32_scan_results.xlsx')
    
    args = parser.parse_args()
    target_drive = args.target_drive
    xlsx_file = args.xlsx_file
    target_exts = [ext.upper() for ext in args.extensions]  # 大文字に揃える

    if args.scan:
        read_raw_data(target_drive, target_exts, xlsx_file)
        lookup_path(xlsx_file)
    
    elif args.restore:
        if not os.path.exists(xlsx_file):
            sys.exit(f"エクセルファイル: {xlsx_file}が見つかりません！")
        salvage_file(xlsx_file)

